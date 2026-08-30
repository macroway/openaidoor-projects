"""
Excel 生成服务。
- 使用 openpyxl 生成双 Sheet Excel
- Sheet 1「询价整理表」：9 列
- Sheet 2「需要人工确认的问题」：3 列
- 设置列宽、表头样式、紧急度单元格颜色
"""
import io
import logging
from typing import List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Sheet 1 列定义：(表头, 字段key, 列宽)
SHEET1_COLUMNS = [
    ("客户名称", "customer_name", 18),
    ("商品", "product", 16),
    ("规格", "spec", 24),
    ("数量", "quantity", 12),
    ("期望价格", "target_price", 14),
    ("交期和到货地", "delivery", 28),
    ("紧急度", "urgency", 10),
    ("缺失信息", "missing_info", 20),
    ("原文依据", "source_text", 30),
]

# Sheet 2 列定义
SHEET2_COLUMNS = [
    ("客户名称", "customer_name", 18),
    ("问题所在", "issue", 30),
    ("应该向客户追问什么", "suggestion", 40),
]

# 样式
HEADER_FONT = Font(name="PingFang SC", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

CELL_FONT = Font(name="PingFang SC", size=10)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)

# 紧急度颜色
URGENCY_FILLS = {
    "普通": PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"),
    "较急": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    "紧急": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
}

URGENCY_FONTS = {
    "普通": Font(name="PingFang SC", size=10, color="6B7280"),
    "较急": Font(name="PingFang SC", size=10, color="D97706"),
    "紧急": Font(name="PingFang SC", size=10, color="DC2626"),
}


def build_excel(
    records: List[Dict[str, Any]],
    confirm_items: List[Dict[str, Any]],
) -> bytes:
    """
    生成双 Sheet Excel 文件。

    Args:
        records: 询价记录列表（已编辑的最终版本）
        confirm_items: 人工确认问题列表

    Returns:
        Excel 文件的 bytes
    """
    wb = Workbook()

    # Sheet 1: 询价整理表
    ws1 = wb.active
    ws1.title = "询价整理表"
    _write_sheet(ws1, SHEET1_COLUMNS, records, is_inquiry=True)

    # Sheet 2: 人工确认问题
    ws2 = wb.create_sheet(title="需要人工确认的问题")
    _write_sheet(ws2, SHEET2_COLUMNS, confirm_items, is_inquiry=False)

    # 输出到 bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    logger.info(f"Excel 生成完成: {len(records)} 条记录, {len(confirm_items)} 个确认问题")
    return output.getvalue()


def _write_sheet(ws, columns, data, is_inquiry=False):
    """写入一个 Sheet 的数据。"""
    # 写表头
    for col_idx, (header, _, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 写数据行
    for row_idx, record in enumerate(data, 2):
        for col_idx, (header, field, _) in enumerate(columns, 1):
            value = record.get(field, "")
            if value is None:
                value = ""

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGN
            cell.border = THIN_BORDER

            # 紧急度列特殊样式
            if field == "urgency" and is_inquiry:
                urgency_val = str(value)
                cell.fill = URGENCY_FILLS.get(urgency_val, URGENCY_FILLS["普通"])
                cell.font = URGENCY_FONTS.get(urgency_val, URGENCY_FONTS["普通"])
                cell.alignment = Alignment(horizontal="center", vertical="top")

            # 数量列加粗
            if field == "quantity" and is_inquiry:
                cell.font = Font(name="PingFang SC", size=11, bold=True, color="059669")

    # 冻结首行
    ws.freeze_panes = "A2"

    # 设置行高
    ws.row_dimensions[1].height = 30
    for row_idx in range(2, len(data) + 2):
        ws.row_dimensions[row_idx].height = 40
